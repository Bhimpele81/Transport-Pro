"""
Bus Route Optimizer - Elbow Lane Day Camp
==========================================

* Geocodes every street address via Google Maps API (primary) with
  OpenStreetMap Nominatim as fallback
* Clusters students by TRUE address-level proximity - ZIP codes ignored
* Gets real driving times via Google Directions API (falls back to OSRM)
* Compass-aware clustering: stops in different compass directions from camp
  are never assigned to the same vehicle (prevents Ambler+Chalfont zigzags)
* Eliminates empty vehicles; warns about under-filled ones in spreadsheet
* Outputs a formatted Excel workbook in Elbow Lane brand colors

Dependencies: pip install openpyxl
"""

import argparse, csv, io, json, math, os, re, time, urllib.parse, urllib.request
from dataclasses import dataclass, field
from typing import Optional, Callable
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# -- Brand / tuneable constants -----------------------------------------------
GOOGLE_MAPS_KEY = os.environ.get("GOOGLE_MAPS_KEY", "")
CAMP_ADDRESS    = "828 Elbow Lane, Warrington, PA 18976"
CAMP_COORDS     = (40.2454, -75.1407)
GEOCACHE_FILE   = "geocache.json"
ROUTECACHE_FILE = "routecache.json"
COORD_OVERRIDES_FILE = "coord_overrides.json"

NEIGHBOR_MI   = 1.5
MIN_UTIL      = 0.60
ROAD_FACTOR   = 1.35
MPH_SUBURBAN  = 30.0

# Maximum angular spread (degrees) allowed between stops on the same vehicle.
# Ambler is 217 deg (SW) from camp, Chalfont is 309 deg (NW) = 92 deg apart.
# Setting this to 75 keeps directionally-incompatible stops off the same bus.
MAX_BEARING_SPREAD_DEG = 75.0

# Distance tier width for sequencing - stops within this range of each other
# in distance-from-camp are grouped and sorted by nearest-neighbor
TIER_WIDTH_MI = 1.0

# -- Coordinate overrides -----------------------------------------------------
def _load_overrides() -> dict:
    return _load_json(COORD_OVERRIDES_FILE)

def add_coord_override(address: str, lat: float, lon: float) -> None:
    overrides = _load_overrides()
    key = address.strip().lower()
    overrides[key] = [lat, lon]
    _save_json(COORD_OVERRIDES_FILE, overrides)
    cache = _load_json(GEOCACHE_FILE)
    old_coords = cache.get(key)
    cache[key] = [lat, lon]
    _save_json(GEOCACHE_FILE, cache)
    if old_coords:
        rcache = _load_json(ROUTECACHE_FILE)
        bad_prefix = f"{old_coords[0]:.5f},{old_coords[1]:.5f}"
        stale = [k for k in rcache if bad_prefix in k]
        for k in stale:
            del rcache[k]
        if stale:
            _save_json(ROUTECACHE_FILE, rcache)

# -- Excel colour palette ------------------------------------------------------
BRAND_COLOR = "6D1F2F"
BRAND_LIGHT = "F5E6E9"
LIGHT_GRAY  = "F2F2F2"
ORANGE_FILL = "FFD966"
GREEN_FILL  = "E2EFDA"
WHITE       = "FFFFFF"
DARK_TEXT   = "1A1A1A"
MED_SIDE  = Side(style="medium", color=BRAND_COLOR)
THIN_SIDE = Side(style="thin",   color="CCCCCC")

# -----------------------------------------------------------------------------
# Data structures
# -----------------------------------------------------------------------------

@dataclass
class Student:
    idx: int
    last: str
    first: str
    address: str
    city: str
    zip_code: str
    lat: float = 0.0
    lon: float = 0.0
    geocoded: bool = False

    @property
    def full_address(self) -> str:
        return f"{self.address}, {self.city}, PA {self.zip_code}"

@dataclass
class Stop:
    address: str
    riders: list = field(default_factory=list)
    drive_time: str = ""
    lat: float = 0.0
    lon: float = 0.0

    @property
    def rider_count(self) -> int:
        return len(self.riders)

    @property
    def rider_names(self) -> str:
        freq: dict = {}
        for s in self.riders:
            freq[s.last] = freq.get(s.last, 0) + 1
        ctr: dict = {}
        names = []
        for s in self.riders:
            if freq[s.last] > 1:
                ctr[s.last] = ctr.get(s.last, 0) + 1
                names.append(f"{s.last}{ctr[s.last]}")
            else:
                names.append(s.last)
        return ", ".join(names)

@dataclass
class Vehicle:
    name: str
    start_address: str
    capacity: int
    stops: list = field(default_factory=list)
    total_time: str = ""
    total_distance: str = ""
    under_threshold: bool = False
    start_lat: float = 0.0
    start_lon: float = 0.0
    camp_lat: float = 0.0
    camp_lon: float = 0.0
    last_leg_mins: int = 0  # drive time from last stop to camp (Google Maps)

    @property
    def rider_count(self) -> int:
        return sum(s.rider_count for s in self.stops)

    @property
    def stop_count(self) -> int:
        return len(self.stops)

    @property
    def utilization_pct(self) -> int:
        return round(self.rider_count / self.capacity * 100) if self.capacity else 0

    @property
    def corridor(self) -> str:
        cities, seen = [], set()
        for stop in self.stops:
            parts = stop.address.split(",")
            city = parts[1].strip() if len(parts) >= 3 else ""
            if city and city not in seen:
                seen.add(city)
                cities.append(city)
        start_city = (self.start_address.split(",")[1].strip()
                      if "," in self.start_address else "")
        return f"{start_city} -> " + " -> ".join(cities[:3]) if cities else start_city

# -----------------------------------------------------------------------------
# Geometry helpers
# -----------------------------------------------------------------------------

def haversine_mi(lat1, lon1, lat2, lon2) -> float:
    R = 3958.8
    p1, p2 = math.radians(lat1), math.radians(lat2)
    a = (math.sin(math.radians(lat2-lat1)/2)**2
         + math.cos(p1)*math.cos(p2)*math.sin(math.radians(lon2-lon1)/2)**2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))

def bearing_deg(from_lat, from_lon, to_lat, to_lon) -> float:
    """Compass bearing in degrees (0=N, 90=E, 180=S, 270=W)."""
    lat1 = math.radians(from_lat)
    lat2 = math.radians(to_lat)
    dlon = math.radians(to_lon - from_lon)
    x = math.sin(dlon) * math.cos(lat2)
    y = math.cos(lat1) * math.sin(lat2) - math.sin(lat1) * math.cos(lat2) * math.cos(dlon)
    return (math.degrees(math.atan2(x, y)) + 360) % 360

def _bearing_spread(bearings: list) -> float:
    """Minimum arc in degrees that contains all bearings (0-360)."""
    if len(bearings) <= 1:
        return 0.0
    sorted_b = sorted(bearings)
    max_gap = 0.0
    for i in range(len(sorted_b)):
        gap = (sorted_b[(i+1) % len(sorted_b)] - sorted_b[i]) % 360
        max_gap = max(max_gap, gap)
    return 360.0 - max_gap

def _bearing_compatible(existing: list, new_bearing: float) -> bool:
    """True if adding new_bearing keeps total spread within MAX_BEARING_SPREAD_DEG."""
    if not existing:
        return True
    return _bearing_spread(existing + [new_bearing]) <= MAX_BEARING_SPREAD_DEG

def centroid(units: list) -> tuple:
    lats = [u[0].lat for u in units]
    lons = [u[0].lon for u in units]
    return sum(lats)/len(lats), sum(lons)/len(lons)

# -----------------------------------------------------------------------------
# JSON cache helpers
# -----------------------------------------------------------------------------

def _load_json(path: str) -> dict:
    if os.path.exists(path):
        try:
            with open(path) as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_json(path: str, data: dict) -> None:
    try:
        with open(path, "w") as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass

# -----------------------------------------------------------------------------
# Driving times
# -----------------------------------------------------------------------------

def _fallback_minutes(lat1, lon1, lat2, lon2) -> float:
    road_mi = haversine_mi(lat1, lon1, lat2, lon2) * ROAD_FACTOR
    return (road_mi / MPH_SUBURBAN) * 60.0

def driving_minutes(lat1, lon1, lat2, lon2, cache: dict) -> float:
    key = f"{lat1:.5f},{lon1:.5f}|{lat2:.5f},{lon2:.5f}"
    if key in cache:
        return cache[key]

    if GOOGLE_MAPS_KEY:
        try:
            params = urllib.parse.urlencode({
                "origin": f"{lat1},{lon1}",
                "destination": f"{lat2},{lon2}",
                "mode": "driving",
                "key": GOOGLE_MAPS_KEY,
            })
            url = f"https://maps.googleapis.com/maps/api/directions/json?{params}"
            req = urllib.request.Request(url,
                headers={"User-Agent": "ElbowLaneCampBusRouter/1.0"})
            with urllib.request.urlopen(req, timeout=8) as resp:
                data = json.loads(resp.read().decode())
            if data.get("status") == "OK":
                mins = data["routes"][0]["legs"][0]["duration"]["value"] / 60.0
                cache[key] = mins
                _save_json(ROUTECACHE_FILE, cache)
                return mins
        except Exception:
            pass

    try:
        url = (f"http://router.project-osrm.org/route/v1/driving/"
               f"{lon1:.6f},{lat1:.6f};{lon2:.6f},{lat2:.6f}?overview=false")
        req = urllib.request.Request(url,
            headers={"User-Agent": "ElbowLaneCampBusRouter/1.0"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode())
        if data.get("code") == "Ok":
            mins = data["routes"][0]["duration"] / 60.0
            cache[key] = mins
            _save_json(ROUTECACHE_FILE, cache)
            return mins
    except Exception:
        pass

    mins = _fallback_minutes(lat1, lon1, lat2, lon2)
    cache[key] = mins
    _save_json(ROUTECACHE_FILE, cache)
    return mins

def route_leg_times(coord_seq: list, progress_cb=None) -> list:
    """
    Returns driving minutes for each leg in coord_seq.
    coord_seq = [garage, stop1, stop2, ..., stopN, camp]
    result[0] = garage->stop1 (deadhead)
    result[i] = stop(i-1)->stop(i) for i>=1
    result[N] = stopN->camp
    """
    cache = _load_json(ROUTECACHE_FILE)
    times = []
    for i in range(len(coord_seq) - 1):
        lat1, lon1 = coord_seq[i]
        lat2, lon2 = coord_seq[i+1]
        times.append(driving_minutes(lat1, lon1, lat2, lon2, cache))
    return times

# -----------------------------------------------------------------------------
# Geocoding
# -----------------------------------------------------------------------------

PA_LAT = (39.7, 42.3)
PA_LON = (-80.5, -74.7)

def _in_pa(lat: float, lon: float) -> bool:
    return PA_LAT[0] <= lat <= PA_LAT[1] and PA_LON[0] <= lon <= PA_LON[1]

MAX_ZIP_DEVIATION_MI = 3.0

ZIP_CENTROIDS = {
    "18901": (40.310, -75.130), "18902": (40.281, -75.095),
    "18914": (40.286, -75.207), "18929": (40.250, -75.084),
    "18954": (40.217, -74.999), "18974": (40.231, -75.062),
    "18976": (40.245, -75.141), "19002": (40.157, -75.228),
    "19025": (40.139, -75.177), "19040": (40.181, -75.106),
    "19044": (40.190, -75.126), "19090": (40.149, -75.120),
    "19446": (40.241, -75.284),
}

def _extract_zip5(address: str) -> str:
    m_zip = re.search(r"\b(\d{5})(?:-\d{4})?\b", address)
    return m_zip.group(1) if m_zip else ""

def _result_near_zip(lat: float, lon: float, zip5: str) -> bool:
    if zip5 not in ZIP_CENTROIDS:
        return True
    clat, clon = ZIP_CENTROIDS[zip5]
    return haversine_mi(lat, lon, clat, clon) <= MAX_ZIP_DEVIATION_MI

def _geocode_one(address: str, cache: dict, progress_cb=None) -> tuple:
    key = address.strip().lower()
    overrides = _load_overrides()
    if key in overrides:
        lat, lon = float(overrides[key][0]), float(overrides[key][1])
        cache[key] = [lat, lon]
        return lat, lon

    if key in cache:
        cached = tuple(cache[key])
        zip5_check = _extract_zip5(address)
        if (_in_pa(*cached) and cached != CAMP_COORDS
                and _result_near_zip(*cached, zip5_check)):
            return cached
        if progress_cb:
            progress_cb(f"  Clearing bad cached coord for {address}: {cached}")
        rcache = _load_json(ROUTECACHE_FILE)
        bad_prefix = f"{cached[0]:.5f},{cached[1]:.5f}"
        stale = [k for k in rcache if bad_prefix in k]
        for k in stale:
            del rcache[k]
        if stale:
            _save_json(ROUTECACHE_FILE, rcache)
        del cache[key]

    if progress_cb:
        progress_cb(f"  Geocoding: {address}")

    zip5  = _extract_zip5(address)
    parts = [p.strip() for p in address.split(",")]
    street = parts[0] if parts else address
    city   = parts[1] if len(parts) > 1 else ""

    def _validate(lat, lon) -> bool:
        return _in_pa(lat, lon) and _result_near_zip(lat, lon, zip5)

    result = None

    if GOOGLE_MAPS_KEY:
        try:
            params = urllib.parse.urlencode({
                "address": address,
                "components": "country:US|administrative_area:PA",
                "key": GOOGLE_MAPS_KEY,
            })
            url = f"https://maps.googleapis.com/maps/api/geocode/json?{params}"
            req = urllib.request.Request(url,
                headers={"User-Agent": "ElbowLaneCampBusRouter/1.0"})
            with urllib.request.urlopen(req, timeout=8) as resp:
                data = json.loads(resp.read().decode())
            if data.get("status") == "OK" and data.get("results"):
                loc = data["results"][0]["geometry"]["location"]
                lat, lon = float(loc["lat"]), float(loc["lng"])
                if _validate(lat, lon):
                    result = (lat, lon)
                elif progress_cb:
                    progress_cb(f"  Google result out of PA bounds for {address}")
        except Exception as e:
            if progress_cb:
                progress_cb(f"  Google geocode error: {e}")

    if result is None:
        try:
            params = urllib.parse.urlencode({
                "street": street, "city": city,
                "state": "Pennsylvania", "country": "United States",
                "format": "json", "limit": 3, "addressdetails": 0,
            })
            req = urllib.request.Request(
                f"https://nominatim.openstreetmap.org/search?{params}",
                headers={"User-Agent": "ElbowLaneCampBusRouter/1.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                results = json.loads(resp.read().decode())
            time.sleep(1.1)
            for r in results:
                lat, lon = float(r["lat"]), float(r["lon"])
                if _validate(lat, lon):
                    result = (lat, lon)
                    break
        except Exception:
            pass

    if result is None:
        try:
            params = urllib.parse.urlencode({
                "q": address, "format": "json", "limit": 5,
                "addressdetails": 0,
                "viewbox": "-80.5,39.7,-74.7,42.3", "bounded": 1,
            })
            req = urllib.request.Request(
                f"https://nominatim.openstreetmap.org/search?{params}",
                headers={"User-Agent": "ElbowLaneCampBusRouter/1.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                results = json.loads(resp.read().decode())
            time.sleep(1.1)
            for r in results:
                lat, lon = float(r["lat"]), float(r["lon"])
                if _validate(lat, lon):
                    result = (lat, lon)
                    break
        except Exception:
            pass

    if result is None and zip5 in ZIP_CENTROIDS:
        lat, lon = ZIP_CENTROIDS[zip5]
        if progress_cb:
            progress_cb(f"  Using ZIP centroid for '{address}' - may be approximate")
        result = (lat, lon)

    if result is not None:
        lat, lon = result
        cache[key] = [lat, lon]
        _save_json(GEOCACHE_FILE, cache)
        if progress_cb:
            progress_cb(f"  OK {address} -> ({lat:.4f}, {lon:.4f})")
        return lat, lon

    if progress_cb:
        progress_cb(f"  Could not geocode '{address}' - using camp coords as fallback")
    return CAMP_COORDS

def _purge_bad_geocache(cache: dict, addresses: list, progress_cb=None) -> int:
    rcache = _load_json(ROUTECACHE_FILE)
    removed = 0
    for addr in addresses:
        key = addr.strip().lower()
        if key not in cache:
            continue
        cached = tuple(cache[key])
        zip5 = _extract_zip5(addr)
        is_bad = (cached == CAMP_COORDS or not _in_pa(*cached)
                  or not _result_near_zip(*cached, zip5))
        if is_bad:
            bad_prefix = f"{cached[0]:.5f},{cached[1]:.5f}"
            stale_routes = [k for k in rcache if bad_prefix in k]
            for k in stale_routes:
                del rcache[k]
            del cache[key]
            removed += 1
            if progress_cb:
                progress_cb(f"  Purged bad geocache entry: {addr} was {cached}")
    if removed:
        _save_json(GEOCACHE_FILE, cache)
        _save_json(ROUTECACHE_FILE, rcache)
    return removed

def clear_bad_geocache() -> int:
    cache = _load_json(GEOCACHE_FILE)
    bad_keys = [k for k, v in cache.items() if not _in_pa(v[0], v[1])]
    for k in bad_keys:
        del cache[k]
    if bad_keys:
        _save_json(GEOCACHE_FILE, cache)
    return len(bad_keys)

def geocode_all_addresses(addresses: list, progress_cb=None) -> dict:
    cache = _load_json(GEOCACHE_FILE)
    purged = _purge_bad_geocache(cache, addresses, progress_cb)
    if purged and progress_cb:
        progress_cb(f"  Purged {purged} bad geocache entries - will re-geocode")

    def needs_geocode(a: str) -> bool:
        key = a.strip().lower()
        if key not in cache:
            return True
        cached = tuple(cache[key])
        if cached == CAMP_COORDS or not _in_pa(*cached):
            return True
        zip5 = _extract_zip5(a)
        return not _result_near_zip(*cached, zip5)

    bad_addresses = [a for a in addresses if needs_geocode(a)
                     and a.strip().lower() in cache]
    if bad_addresses:
        rcache = _load_json(ROUTECACHE_FILE)
        removed_routes = 0
        for addr in bad_addresses:
            old_key = cache.get(addr.strip().lower())
            if old_key:
                bad_coord_prefix = f"{old_key[0]:.5f},{old_key[1]:.5f}"
                stale = [k for k in rcache if bad_coord_prefix in k]
                for k in stale:
                    del rcache[k]
                removed_routes += 1
        if removed_routes:
            _save_json(ROUTECACHE_FILE, rcache)
            if progress_cb:
                progress_cb(f"  Cleared {removed_routes} stale route-time entries")

    new_count = sum(1 for a in addresses if needs_geocode(a))
    if new_count and progress_cb:
        progress_cb(f"Geocoding {new_count} address(es) via Google Maps...")
    elif progress_cb:
        progress_cb(f"All {len(addresses)} addresses loaded from cache")

    return {a: _geocode_one(a, cache, progress_cb) for a in addresses}

# -----------------------------------------------------------------------------
# CSV parsing
# -----------------------------------------------------------------------------

def parse_students_csv(csv_text: str) -> list:
    if "\n" not in csv_text and len(csv_text) < 300:
        try:
            with open(csv_text, newline="", encoding="utf-8-sig") as f:
                csv_text = f.read()
        except FileNotFoundError:
            pass
    students = []
    reader = csv.DictReader(io.StringIO(csv_text))
    for row in reader:
        row = {k.strip().strip('"'): v.strip().strip('"') for k, v in row.items()}
        idx_key = next((k for k in row if k in ("", "idx", "#")), "")
        try: idx = int(row.get(idx_key, 0))
        except: idx = 0
        last  = (row.get("Last name") or row.get("last_name") or row.get("Last Name") or "")
        first = (row.get("First name") or row.get("first_name") or row.get("First Name") or "")
        addr  = (row.get("Primary family address 1") or row.get("Address") or
                 row.get("address") or row.get("Street") or "")
        city  = (row.get("Primary family city") or row.get("City") or row.get("city") or "")
        zip_  = (row.get("Primary family zip") or row.get("Zip") or row.get("zip")
                 or row.get("ZIP") or row.get("Postal Code") or "")
        if last and addr:
            students.append(Student(idx, last, first, addr, city, zip_))
    return students

# -----------------------------------------------------------------------------
# Vehicle config parsing
# -----------------------------------------------------------------------------

def parse_vehicles_text(text: str) -> list:
    VEH_RE = re.compile(
        r"""^(?:(?:vehicles?\s+[A-Z0-9][-\s,A-Z0-9]*)
        |(?:[A-Z]-[A-Z]\s*(?:\(|$))
        |(?:[A-Z]\s*(?:\(|:|\s+Start))
        |(?:Van\s+[A-Z0-9]))""",
        re.VERBOSE | re.IGNORECASE)

    merged = []
    for raw in text.strip().splitlines():
        s = raw.strip()
        if not s: continue
        if merged and not VEH_RE.match(s):
            merged[-1] += " " + s
        else:
            merged.append(s)

    RNG = re.compile(r"^(?:vehicles?\s*)?([A-Z])[-]([A-Z])(?:\s*\(\d+\s*vehicles?\))?",
                     re.IGNORECASE)
    expanded = []
    for line in merged:
        m = RNG.match(line)
        if m:
            remainder = line[m.end():].strip().lstrip(":")
            for c in range(ord(m.group(1).upper()), ord(m.group(2).upper())+1):
                expanded.append(f"Vehicle {chr(c)}: {remainder}")
        else:
            expanded.append(line)

    vehicles = []
    for line in expanded:
        line = line.strip()
        nm = re.match(r"^((?:Vehicles?|Van)\s+[A-Z0-9]+)", line, re.I)
        if not nm:
            nm = re.match(r"^([A-Z])(?:\s*:|\s+Start)", line, re.I)
        if not nm: continue
        name = re.sub(r"^Vehicles\b", "Vehicle", nm.group(1).strip(), flags=re.I)
        rest = line[nm.end():].strip().lstrip(":").strip()
        sm = re.search(r"Start(?:\s*[&]\s*End)?\s*:?\s*", rest, re.I)
        if not sm: continue
        after = rest[sm.end():]
        ae = re.search(r"\s*[-]?\s*Capacity\b", after, re.I)
        start_addr = (after[:ae.start()].strip() if ae else after.strip()).rstrip(" -,")
        cm = re.search(r"Capacity\s*:?\s*(?:up\s+to\s+)?(\d+)\s*riders?", rest, re.I)
        cap = int(cm.group(1)) if cm else 13
        if name and start_addr:
            vehicles.append({"name": name, "start": start_addr, "capacity": cap})
    return vehicles

# -----------------------------------------------------------------------------
# TSP Sequencing - tier-based camp-directional
# -----------------------------------------------------------------------------

def _sequence_stops_camp_directional(stops: list, camp_lat: float, camp_lon: float,
                                      trip_direction: str = "morning") -> list:
    if len(stops) <= 1:
        return list(stops)

    def d2c(s):
        return haversine_mi(s.lat, s.lon, camp_lat, camp_lon)

    if trip_direction == "afternoon":
        sorted_stops = sorted(stops, key=d2c)
    else:
        sorted_stops = sorted(stops, key=d2c, reverse=True)

    # Group into distance tiers
    tiers = []
    current_tier = [sorted_stops[0]]
    ref_d = d2c(sorted_stops[0])

    for s in sorted_stops[1:]:
        if abs(d2c(s) - ref_d) <= TIER_WIDTH_MI:
            current_tier.append(s)
        else:
            tiers.append(current_tier)
            current_tier = [s]
            ref_d = d2c(s)
    tiers.append(current_tier)

    # Within each tier, use nearest-neighbor
    result = []
    last_lat = sum(s.lat for s in tiers[0]) / len(tiers[0])
    last_lon = sum(s.lon for s in tiers[0]) / len(tiers[0])

    for tier in tiers:
        remaining = list(tier)
        while remaining:
            nearest = min(remaining,
                key=lambda s: haversine_mi(last_lat, last_lon, s.lat, s.lon))
            result.append(nearest)
            last_lat, last_lon = nearest.lat, nearest.lon
            remaining.remove(nearest)

    return result

# -----------------------------------------------------------------------------
# Core routing
# -----------------------------------------------------------------------------

def cluster_and_route(students: list, vehicles: list,
                      progress_cb: Optional[Callable] = None,
                      camp_address: str = None,
                      trip_direction: str = "morning") -> list:
    effective_camp = camp_address or CAMP_ADDRESS

    # -- 1. Geocode -----------------------------------------------------------
    all_addrs = list({s.full_address for s in students}
                     | {v["start"] for v in vehicles}
                     | {effective_camp})
    coords = geocode_all_addresses(all_addrs, progress_cb)
    camp_lat, camp_lon = coords.get(effective_camp, CAMP_COORDS)

    for s in students:
        s.lat, s.lon = coords.get(s.full_address, CAMP_COORDS)
        s.geocoded = (s.lat, s.lon) != CAMP_COORDS

    if progress_cb:
        ok = sum(1 for s in students if s.geocoded)
        progress_cb(f"Geocoded {ok}/{len(students)} addresses")
        for s in [s for s in students if not s.geocoded][:5]:
            progress_cb(f"  Could not geocode: {s.full_address}")

    # -- 2. Group by exact address (families) ---------------------------------
    addr_map: dict = {}
    for s in students:
        addr_map.setdefault(s.address.lower().strip(), []).append(s)
    family_units = list(addr_map.values())

    def uc(u): return u[0].lat, u[0].lon
    def d2c(u): return haversine_mi(*uc(u), camp_lat, camp_lon)

    # -- 3. Geographic proximity clustering -----------------------------------
    clusters: list = []
    for unit in family_units:
        ulat, ulon = uc(unit)
        best_ci, best_d = None, float("inf")
        for ci, cluster in enumerate(clusters):
            for eu in cluster:
                d = haversine_mi(ulat, ulon, *uc(eu))
                if d <= NEIGHBOR_MI and d < best_d:
                    best_d, best_ci = d, ci
        if best_ci is not None:
            clusters[best_ci].append(unit)
        else:
            clusters.append([unit])

    if progress_cb:
        progress_cb(f"Formed {len(clusters)} geographic clusters "
                    f"(<={NEIGHBOR_MI} mi between actual house coordinates)")

    # -- 4. Build vehicle objects ---------------------------------------------
    def cl_size(cl): return sum(len(u) for u in cl)

    veh_objects = []
    for v in vehicles:
        lat, lon = coords.get(v["start"], CAMP_COORDS)
        veh_objects.append(Vehicle(
            name=v["name"], start_address=v["start"],
            capacity=v["capacity"], start_lat=lat, start_lon=lon))

    # -- 5. Compass-aware cluster assignment ----------------------------------
    max_cap = max(v["capacity"] for v in vehicles)

    assignable: list = []
    for cl in clusters:
        if cl_size(cl) <= max_cap:
            assignable.append(cl)
        else:
            units_sorted = sorted(cl, key=d2c, reverse=True)
            chunk, chunk_n = [], 0
            for u in units_sorted:
                if chunk_n + len(u) > max_cap and chunk:
                    assignable.append(chunk); chunk, chunk_n = [], 0
                chunk.append(u); chunk_n += len(u)
            if chunk: assignable.append(chunk)

    assignable.sort(
        key=lambda cl: haversine_mi(*centroid(cl), camp_lat, camp_lon),
        reverse=True)

    assignments = [[] for _ in veh_objects]
    counts = [0] * len(veh_objects)
    veh_bearings = [[] for _ in veh_objects]

    def _cl_bearing(cl):
        clat, clon = centroid(cl)
        return bearing_deg(camp_lat, camp_lon, clat, clon)

    def _best_vehicle(cl, sz):
        cl_b = _cl_bearing(cl)
        clat, clon = centroid(cl)
        compatible = []
        fallback = []
        for vi in range(len(veh_objects)):
            if veh_objects[vi].capacity - counts[vi] < sz:
                continue
            geo = haversine_mi(clat, clon,
                               veh_objects[vi].start_lat,
                               veh_objects[vi].start_lon)
            if _bearing_compatible(veh_bearings[vi], cl_b):
                compatible.append((geo, vi))
            else:
                fallback.append((geo, vi))
        compatible.sort()
        fallback.sort()
        if compatible:
            return compatible[0][1]
        if fallback:
            return fallback[0][1]
        return min(range(len(veh_objects)), key=lambda i: counts[i])

    for cl in assignable:
        sz = cl_size(cl)
        vi = _best_vehicle(cl, sz)
        cl_b = _cl_bearing(cl)
        for u in cl:
            assignments[vi].append(u)
        counts[vi] += sz
        veh_bearings[vi].append(cl_b)

    # -- 5b. Consolidation ----------------------------------------------------
    changed, passes = True, 0
    while changed and passes < 30:
        changed, passes = False, passes + 1

        def effective_threshold(vi):
            cap = veh_objects[vi].capacity
            if cap <= 6:  return 0.40
            if cap <= 9:  return 0.50
            return MIN_UTIL

        under = sorted(
            [vi for vi in range(len(veh_objects))
             if counts[vi] > 0 and
             counts[vi] / veh_objects[vi].capacity < effective_threshold(vi)],
            key=lambda vi: counts[vi])

        if not under: break

        vi_src = under[0]
        units_src = assignments[vi_src][:]
        total_src = counts[vi_src]

        if not units_src:
            changed = True; continue

        src_lat, src_lon = centroid(units_src)
        src_bearing = bearing_deg(camp_lat, camp_lon, src_lat, src_lon)

        full_dest, full_dist = None, float("inf")
        for vi_dst in range(len(veh_objects)):
            if vi_dst == vi_src: continue
            if veh_objects[vi_dst].capacity - counts[vi_dst] < total_src: continue
            if not _bearing_compatible(veh_bearings[vi_dst], src_bearing):
                continue
            if assignments[vi_dst]:
                dlat, dlon = centroid(assignments[vi_dst])
            else:
                dlat, dlon = veh_objects[vi_dst].start_lat, veh_objects[vi_dst].start_lon
            d = haversine_mi(src_lat, src_lon, dlat, dlon)
            if d < full_dist: full_dist, full_dest = d, vi_dst

        MAX_MERGE_MI = 5.0
        if full_dest is not None and full_dist <= MAX_MERGE_MI:
            for u in units_src: assignments[full_dest].append(u)
            counts[full_dest] += total_src
            veh_bearings[full_dest].append(src_bearing)
            assignments[vi_src] = []; counts[vi_src] = 0
            veh_bearings[vi_src] = []
            if progress_cb:
                progress_cb(f"  Merged {veh_objects[vi_src].name} ({total_src}) -> "
                            f"{veh_objects[full_dest].name} "
                            f"({counts[full_dest]}/{veh_objects[full_dest].capacity})")
            changed = True; continue

        remaining = list(units_src)
        sub_groups = []
        while remaining:
            group = [remaining[0]]
            remaining.pop(0)
            changed_inner = True
            while changed_inner:
                changed_inner = False
                for unit in list(remaining):
                    ulat, ulon = uc(unit)
                    if any(haversine_mi(ulat, ulon, *uc(g)) <= NEIGHBOR_MI
                           for g in group):
                        group.append(unit)
                        remaining.remove(unit)
                        changed_inner = True
            sub_groups.append(group)

        sub_groups.sort(key=lambda g: sum(len(u) for u in g), reverse=True)

        moved = False
        for group in sub_groups:
            group_size = sum(len(u) for u in group)
            glat = sum(uc(u)[0] for u in group) / len(group)
            glon = sum(uc(u)[1] for u in group) / len(group)
            g_bearing = bearing_deg(camp_lat, camp_lon, glat, glon)
            MAX_SCATTER_MI = 2.5
            best_vi, best_d = None, float("inf")

            for vi_dst in range(len(veh_objects)):
                if vi_dst == vi_src: continue
                if veh_objects[vi_dst].capacity - counts[vi_dst] < group_size: continue
                if not _bearing_compatible(veh_bearings[vi_dst], g_bearing):
                    continue
                if assignments[vi_dst]:
                    d = min(haversine_mi(glat, glon, *uc(eu))
                            for eu in assignments[vi_dst])
                    if d > MAX_SCATTER_MI: continue
                else:
                    d = haversine_mi(glat, glon,
                                     veh_objects[vi_dst].start_lat,
                                     veh_objects[vi_dst].start_lon)
                    if d > MAX_SCATTER_MI * 2: continue
                if d < best_d: best_d, best_vi = d, vi_dst

            if best_vi is not None:
                for unit in group:
                    assignments[best_vi].append(unit)
                    assignments[vi_src].remove(unit)
                counts[best_vi] += group_size
                counts[vi_src]  -= group_size
                veh_bearings[best_vi].append(g_bearing)
                moved = True
                names = ", ".join(u[0].last for u in group)
                if progress_cb:
                    progress_cb(f"  Moved [{names}] ({group_size}) "
                                f"{veh_objects[vi_src].name} -> "
                                f"{veh_objects[best_vi].name}")

        if moved: changed = True

    # -- 5c. Prune empty vehicles ---------------------------------------------
    active = [vi for vi in range(len(veh_objects)) if counts[vi] > 0]
    veh_objects  = [veh_objects[vi]  for vi in active]
    assignments  = [assignments[vi]  for vi in active]
    counts       = [counts[vi]       for vi in active]

    if progress_cb:
        progress_cb(f"Active vehicles: {len(veh_objects)}")

    # -- 6. Sequence stops + driving times ------------------------------------
    for vi, veh in enumerate(veh_objects):
        if not assignments[vi]:
            veh.total_time = "---"; veh.total_distance = "---"; continue

        addr_stop: dict = {}
        for unit in assignments[vi]:
            rep = unit[0]; key = rep.full_address.lower().strip()
            if key not in addr_stop:
                addr_stop[key] = Stop(address=rep.full_address,
                                      lat=rep.lat, lon=rep.lon)
            addr_stop[key].riders.extend(unit)

        if progress_cb:
            zero_stops = [s for s in addr_stop.values() if abs(s.lat) < 0.001]
            for zs in zero_stops:
                progress_cb(f"  Stop has no coordinates: {zs.address}")

        sorted_stops = _sequence_stops_camp_directional(
            list(addr_stop.values()), camp_lat, camp_lon, trip_direction)

        # coord sequence: [garage, stop1, ..., stopN, camp]
        coord_seq = ([(veh.start_lat, veh.start_lon)]
                     + [(s.lat, s.lon) for s in sorted_stops]
                     + [(camp_lat, camp_lon)])

        # legs[0] = garage->stop1 (deadhead, not shown)
        # legs[i] = stop(i-1)->stop(i) for i>=1
        # legs[-1] = stopN->camp
        legs = route_leg_times(coord_seq, progress_cb)

        # Stop 1: blank, stop 2+: stop-to-stop time
        for i, stop in enumerate(sorted_stops):
            if i == 0:
                stop.drive_time = ""
            else:
                mins = max(1, round(legs[i]))
                stop.drive_time = f"{mins} min"

        # Store last leg (stopN -> camp) for arrival row display
        veh.last_leg_mins = max(1, round(legs[-1])) if legs else 0

        veh.stops = sorted_stops
        veh.camp_lat = camp_lat
        veh.camp_lon = camp_lon

        # Kids ride time = legs[1:] (excludes garage deadhead)
        kids_legs = legs[1:]
        kids_mins = round(sum(kids_legs))
        if kids_mins >= 60:
            hrs, rem = divmod(kids_mins, 60)
            veh.total_time = f"{hrs} hr {rem} min" if rem else f"{hrs} hr"
        else:
            veh.total_time = f"{kids_mins} min"

        dist = sum(
            haversine_mi(coord_seq[i][0], coord_seq[i][1],
                         coord_seq[i+1][0], coord_seq[i+1][1]) * ROAD_FACTOR
            for i in range(len(coord_seq)-1))
        veh.total_distance = f"{round(dist, 1)} mi"

        cap = veh.capacity
        eff_threshold = 0.50 if cap <= 6 else (0.60 if cap <= 9 else MIN_UTIL)
        veh.under_threshold = (veh.rider_count / cap < eff_threshold if cap else False)

    return veh_objects

# -----------------------------------------------------------------------------
# Excel output
# -----------------------------------------------------------------------------

def _fill(c): return PatternFill("solid", start_color=c, fgColor=c)
def _font(bold=False, size=11, color=DARK_TEXT, italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _bdr(**kw): return Border(**kw)


def build_dashboard(wb: Workbook, vehicles: list,
                    camp_address: str = None, trip_direction: str = "morning"):
    ws = wb.active
    ws.title = "Route Summary"
    for col, w in zip("ABCDEFGH", [18, 54, 10, 10, 9, 16, 13, 10]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "Elbow Lane Day Camp - Vehicle Route Plan"
    c.font = _font(bold=True, size=16, color=WHITE)
    c.fill = _fill(BRAND_COLOR)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    camp_display = camp_address or CAMP_ADDRESS
    direction_label = ("All vehicles depart from" if trip_direction == "afternoon"
                       else "All vehicles finish at")
    c.value = (f"{direction_label}: {camp_display} | "
               "Compass-aware clustering | "
               "Drive times via Google Maps | "
               "Kids Ride Time = first stop to camp")
    c.font = _font(size=9, italic=True, color="444444")
    c.fill = _fill(BRAND_LIGHT)
    c.alignment = _align("center")
    ws.row_dimensions[2].height = 18

    hdrs = ["Vehicle", "Starting Point / Route Corridor",
            "Capacity", "Riders", "Stops", "Kids Ride Time", "Distance", "Utilization"]
    for ci, h in enumerate(hdrs, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = _font(bold=True, size=10, color=WHITE)
        cell.fill = _fill(BRAND_COLOR)
        cell.alignment = _align("center")
        cell.border = _bdr(bottom=MED_SIDE, top=MED_SIDE,
                           left=THIN_SIDE, right=THIN_SIDE)
    ws.row_dimensions[3].height = 18

    total_cap = total_riders = 0
    has_warnings = False

    for ri, veh in enumerate(vehicles):
        row = 4 + ri
        warn = veh.under_threshold
        if warn: has_warnings = True
        bg = _fill(ORANGE_FILL) if warn else (_fill(WHITE) if ri%2==0 else _fill(LIGHT_GRAY))
        util_txt = f"{veh.utilization_pct}%"
        if warn: util_txt += " !"
        vals = [veh.name,
                f"{veh.start_address} | {veh.corridor}",
                veh.capacity, veh.rider_count, veh.stop_count,
                veh.total_time, veh.total_distance, util_txt]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = _font(size=10, bold=warn)
            cell.fill = bg
            cell.alignment = _align("left" if ci == 2 else "center")
            cell.border = _bdr(bottom=THIN_SIDE)
        total_cap    += veh.capacity
        total_riders += veh.rider_count
        ws.row_dimensions[row].height = 15

    tr = 4 + len(vehicles)
    ws.merge_cells(f"A{tr}:B{tr}")
    c = ws[f"A{tr}"]
    c.value = (f"TOTAL ({total_riders} riders / {total_cap} capacity "
               f"/ {len(vehicles)} vehicles)")
    c.font = _font(bold=True, size=10, color=WHITE)
    c.fill = _fill(BRAND_COLOR)
    c.alignment = _align("center")
    for ci, val in [(3, total_cap), (4, f"=SUM(D4:D{tr-1})"),
                    (5, f"=SUM(E4:E{tr-1})"),
                    (6, "---"), (7, "---"), (8, "---")]:
        cell = ws.cell(row=tr, column=ci, value=val)
        cell.font = _font(bold=True, color=WHITE)
        cell.fill = _fill(BRAND_COLOR)
        cell.alignment = _align("center")
    ws.row_dimensions[tr].height = 18

    lr = tr + 2
    ws.merge_cells(f"A{lr}:H{lr}")
    c = ws[f"A{lr}"]
    if has_warnings:
        c.value = ("! Orange rows are below utilization threshold. "
                   "These vehicles serve geographically isolated stops.")
        c.font = _font(size=9, bold=True, color="7B3F00")
        c.fill = _fill(ORANGE_FILL)
    else:
        c.value = ("All vehicles at target capacity | "
                   "Compass-aware clustering keeps stops in same direction | "
                   "Kids Ride Time = first pickup to camp")
        c.font = _font(size=9, italic=True, color="555555")
    c.alignment = _align("left", wrap=True)
    ws.row_dimensions[lr].height = 28


def build_vehicle_sheet(wb: Workbook, veh: Vehicle,
                        camp_address: str = None, trip_direction: str = "morning"):
    ws = wb.create_sheet(title=veh.name)
    for col, w in zip("ABCDE", [9, 40, 28, 12, 22]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"{veh.name} - Route Sheet"
    c.font = _font(bold=True, size=14, color=WHITE)
    c.fill = _fill(BRAND_COLOR)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    c = ws["A2"]
    warn_tag = " ! Below threshold - see dashboard" if veh.under_threshold else ""
    c.value = (f"Start: {veh.start_address} | Cap: {veh.capacity} | "
               f"Riders: {veh.rider_count} ({veh.utilization_pct}%) | "
               f"Kids Ride Time: {veh.total_time}, {veh.total_distance}{warn_tag}")
    c.font = _font(size=9, italic=True,
                   color="7B3F00" if veh.under_threshold else DARK_TEXT)
    c.fill = _fill(ORANGE_FILL if veh.under_threshold else BRAND_LIGHT)
    c.alignment = _align("left")
    ws.row_dimensions[2].height = 16

    for ci, h in enumerate(["Stop #", "Address", "Riders", "# Riders", "Drive Time"], 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font = _font(bold=True, size=10, color=WHITE)
        cell.fill = _fill(BRAND_COLOR)
        cell.alignment = _align("center")
        cell.border = _bdr(bottom=MED_SIDE)
    ws.row_dimensions[3].height = 16

    for ci, val in enumerate([" START", veh.start_address,
                               "Departure point", None, None], 1):
        cell = ws.cell(row=4, column=ci, value=val)
        cell.font = _font(bold=(ci==1), italic=(ci==3), color="555555")
        cell.fill = _fill(LIGHT_GRAY)
        cell.alignment = _align("left" if ci == 2 else "center")
    ws.row_dimensions[4].height = 14

    for si, stop in enumerate(veh.stops):
        row = 5 + si
        bg = _fill(WHITE) if si % 2 == 0 else _fill(LIGHT_GRAY)
        for ci, val in enumerate([si+1, stop.address, stop.rider_names,
                                   stop.rider_count, stop.drive_time], 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = _font(bold=(ci==1), size=10)
            cell.fill = bg
            cell.alignment = _align("center" if ci in (1,4) else "left")
            cell.border = _bdr(bottom=THIN_SIDE)
        ws.row_dimensions[row].height = 14

    arrive = 5 + len(veh.stops)

    # Use Google Maps last-leg time if available, else fall back to estimate
    if veh.last_leg_mins and veh.last_leg_mins > 0:
        final_mins = veh.last_leg_mins
    elif veh.stops:
        last = veh.stops[-1]
        final_mins = max(1, round(_fallback_minutes(last.lat, last.lon, *CAMP_COORDS)))
    else:
        final_mins = 0

    arrive_time = f"{final_mins} min -> ARRIVE" if final_mins else "-> ARRIVE"
    arrive_label = camp_address or CAMP_ADDRESS
    action_word = "DEPART" if trip_direction == "afternoon" else "ARRIVE"

    for ci, val in enumerate([action_word, arrive_label, "---", "---", arrive_time], 1):
        cell = ws.cell(row=arrive, column=ci, value=val)
        cell.font = _font(bold=True, color="006400")
        cell.fill = _fill(GREEN_FILL)
        cell.alignment = _align("left" if ci == 2 else "center")
        cell.border = _bdr(top=MED_SIDE, bottom=MED_SIDE)
    ws.row_dimensions[arrive].height = 16

    tr = arrive + 1
    ws.cell(row=tr, column=4, value="Total Riders:").font = _font(bold=True)
    ws.cell(row=tr, column=4).alignment = _align("right")
    cell = ws.cell(row=tr, column=5, value=f"=SUM(D5:D{arrive-1})")
    cell.font = _font(bold=True); cell.alignment = _align("center")

    nr = tr + 1
    ws.merge_cells(f"A{nr}:E{nr}")
    note = ws[f"A{nr}"]
    note.value = ("Drive times via Google Maps | "
                  "Compass-aware clustering - stops in same direction from camp | "
                  "Kids Ride Time = first pickup to camp (excludes garage deadhead)")
    note.font = _font(size=8, italic=True, color="777777")
    note.alignment = _align("left")

# -----------------------------------------------------------------------------
# Public API
# -----------------------------------------------------------------------------

def generate_routes(
    csv_text: str,
    vehicles_text: str,
    output_path: str = "bus_routes_output.xlsx",
    route_data: Optional[list] = None,
    progress_cb: Optional[Callable] = None,
    camp_address: str = None,
    trip_direction: str = "morning",
) -> str:
    students = parse_students_csv(csv_text)
    if not students:
        raise ValueError("No students parsed. Check CSV format.")

    vehicle_configs = parse_vehicles_text(vehicles_text)
    if not vehicle_configs:
        raise ValueError("No vehicles parsed. Check fleet configuration format.")

    if progress_cb:
        unique = len({s.full_address for s in students})
        progress_cb(f"Loaded {len(students)} students across {unique} addresses, "
                    f"{len(vehicle_configs)} vehicles")

    if route_data:
        vehicles = _apply_ai_routes(vehicle_configs, route_data)
    else:
        vehicles = cluster_and_route(students, vehicle_configs, progress_cb,
                                     camp_address=camp_address,
                                     trip_direction=trip_direction)

    wb = Workbook()
    build_dashboard(wb, vehicles, camp_address=camp_address, trip_direction=trip_direction)
    for veh in vehicles:
        build_vehicle_sheet(wb, veh, camp_address=camp_address, trip_direction=trip_direction)

    wb.save(output_path)
    if progress_cb:
        progress_cb(f"Saved: {output_path}")
    return output_path


def _apply_ai_routes(vehicle_configs, route_data) -> list:
    vehicles = []
    for rd in route_data:
        veh = Vehicle(name=rd.get("vehicle_name","?"),
                      start_address=rd.get("start_address",""),
                      capacity=rd.get("capacity", 13),
                      total_time=rd.get("total_time","---"),
                      total_distance=rd.get("total_distance","---"))
        for sd in rd.get("stops", []):
            stop = Stop(address=sd.get("address",""))
            for name in sd.get("rider_names", []):
                stop.riders.append(Student(0, name, "", "", "", ""))
            stop.drive_time = sd.get("drive_time","")
            veh.stops.append(stop)
        vehicles.append(veh)
    return vehicles

# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Elbow Lane Camp Bus Router")
    parser.add_argument("--csv", required=True)
    parser.add_argument("--vehicles", required=True)
    parser.add_argument("--output", default="bus_routes_output.xlsx")
    parser.add_argument("--route-json", default=None)
    args = parser.parse_args()

    with open(args.csv, encoding="utf-8-sig") as f:
        csv_text = f.read()
    try:
        with open(args.vehicles) as f:
            vehicles_text = f.read()
    except (FileNotFoundError, IsADirectoryError):
        vehicles_text = args.vehicles.replace("\\n", "\n")

    route_data = None
    if args.route_json:
        with open(args.route_json) as f:
            route_data = json.load(f)

    generate_routes(csv_text, vehicles_text, args.output, route_data, print)

if __name__ == "__main__":
    main()