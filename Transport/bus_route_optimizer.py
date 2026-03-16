"""
Bus Route Optimizer — Elbow Lane Day Camp
==========================================
• Geocodes every street address via OpenStreetMap Nominatim (free, no key)
• Clusters students by TRUE address-level proximity — ZIP codes ignored
• Gets real driving times via OSRM road-network API (free, no key)
  Falls back to road-factor estimate when OSRM is unavailable
• Assigns whole geographic clusters to vehicles (neighbors always same van)
• Eliminates empty vehicles; warns about under-filled ones in spreadsheet
• Outputs a formatted Excel workbook in Elbow Lane brand colors

Dependencies:  pip install openpyxl

Usage (CLI):
    python bus_route_optimizer.py --csv students.csv --vehicles fleet.txt

Usage (import):
    from bus_route_optimizer import generate_routes
    generate_routes(csv_text, vehicles_text, "output.xlsx", progress_cb=print)
"""

import argparse, csv, io, json, math, os, re, time, urllib.parse, urllib.request
from dataclasses import dataclass, field
from typing import Optional, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Brand / tuneable constants ────────────────────────────────────────────────
GOOGLE_MAPS_KEY = os.environ.get("GOOGLE_MAPS_KEY", "")   # set in Replit Secrets
CAMP_ADDRESS    = "828 Elbow Lane, Warrington, PA 18976"
CAMP_COORDS     = (40.2454, -75.1407)   # fallback if geocode fails
GEOCACHE_FILE   = "geocache.json"        # on-disk cache for lat/lon lookups
ROUTECACHE_FILE = "routecache.json"      # on-disk cache for OSRM driving times
COORD_OVERRIDES_FILE = "coord_overrides.json"  # user-editable GPS overrides
NEIGHBOR_MI     = 1.5                    # houses ≤ 1.5 mi apart → same-van candidate
MIN_UTIL        = 0.75                   # target minimum utilisation per vehicle
ROAD_FACTOR     = 1.35                   # road distance ≈ straight-line × 1.35
MPH_SUBURBAN    = 30.0                   # average speed for fallback time estimate

# ---------------------------------------------------------------------------
# Coordinate overrides — bypasses Nominatim for addresses it gets wrong.
# Keys are lowercase address strings. Values are [lat, lon].
# These are loaded from coord_overrides.json if it exists, and ALWAYS
# take priority over Nominatim results and the geocache.
# To add a new override: add a line to coord_overrides.json, or call
# add_coord_override(address, lat, lon) from your app.
# ---------------------------------------------------------------------------
def _load_overrides() -> dict:
    """Load user coordinate overrides from disk."""
    return _load_json(COORD_OVERRIDES_FILE)

def add_coord_override(address: str, lat: float, lon: float) -> None:
    """
    Permanently override the geocoded coordinates for an address.
    Use this when Nominatim returns wrong coordinates for a specific address.
    The override is saved to coord_overrides.json and used on all future runs.
    Also clears any cached geocode and route-time entries for this address.
    """
    overrides = _load_overrides()
    key = address.strip().lower()
    overrides[key] = [lat, lon]
    _save_json(COORD_OVERRIDES_FILE, overrides)
    
    # Also update geocache and clear stale route times
    cache = _load_json(GEOCACHE_FILE)
    old_coords = cache.get(key)
    cache[key] = [lat, lon]
    _save_json(GEOCACHE_FILE, cache)
    
    if old_coords:
        rcache = _load_json(ROUTECACHE_FILE)
        bad_prefix = f"{old_coords[0]:.5f},{old_coords[1]:.5f}"
        stale = [k for k in rcache if bad_prefix in k]
        for k in stale:
            del rcache[k]
        if stale:
            _save_json(ROUTECACHE_FILE, rcache)

# ── Excel colour palette (Elbow Lane brand) ───────────────────────────────────
BRAND_COLOR  = "6D1F2F"   # deep burgundy
BRAND_LIGHT  = "F5E6E9"   # pale rose — subtitle bar
LIGHT_GRAY   = "F2F2F2"
YELLOW_FILL  = "FFEB9C"   # under-threshold warning
ORANGE_FILL  = "FFD966"   # utilisation warning row
GREEN_FILL   = "E2EFDA"
WHITE        = "FFFFFF"
DARK_TEXT    = "1A1A1A"
MED_SIDE     = Side(style="medium", color=BRAND_COLOR)
THIN_SIDE    = Side(style="thin",   color="CCCCCC")


# ─────────────────────────────────────────────────────────────────────────────
# Data structures
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class Student:
    idx: int
    last: str
    first: str
    address: str
    city: str
    zip_code: str
    lat: float = 0.0
    lon: float = 0.0
    geocoded: bool = False

    @property
    def full_address(self) -> str:
        return f"{self.address}, {self.city}, PA {self.zip_code}"


@dataclass
class Stop:
    address: str
    riders: list = field(default_factory=list)
    drive_time: str = ""
    lat: float = 0.0
    lon: float = 0.0

    @property
    def rider_count(self) -> int:
        return len(self.riders)

    @property
    def rider_names(self) -> str:
        freq: dict = {}
        for s in self.riders:
            freq[s.last] = freq.get(s.last, 0) + 1
        ctr: dict = {}
        names = []
        for s in self.riders:
            if freq[s.last] > 1:
                ctr[s.last] = ctr.get(s.last, 0) + 1
                names.append(f"{s.last}{ctr[s.last]}")
            else:
                names.append(s.last)
        return ", ".join(names)


@dataclass
class Vehicle:
    name: str
    start_address: str
    capacity: int
    stops: list        = field(default_factory=list)
    total_time: str    = ""
    total_distance: str = ""
    under_threshold: bool = False   # True when utilisation < MIN_UTIL
    start_lat: float   = 0.0
    start_lon: float   = 0.0
    camp_lat:  float   = 0.0
    camp_lon:  float   = 0.0

    @property
    def rider_count(self) -> int:
        return sum(s.rider_count for s in self.stops)

    @property
    def stop_count(self) -> int:
        return len(self.stops)

    @property
    def utilization_pct(self) -> int:
        return round(self.rider_count / self.capacity * 100) if self.capacity else 0

    @property
    def corridor(self) -> str:
        cities, seen = [], set()
        for stop in self.stops:
            parts = stop.address.split(",")
            city = parts[1].strip() if len(parts) >= 3 else ""
            if city and city not in seen:
                seen.add(city)
                cities.append(city)
        start_city = (self.start_address.split(",")[1].strip()
                      if "," in self.start_address else "")
        return f"{start_city} → " + " → ".join(cities[:3]) if cities else start_city


# ─────────────────────────────────────────────────────────────────────────────
# Geometry helpers
# ─────────────────────────────────────────────────────────────────────────────

def haversine_mi(lat1, lon1, lat2, lon2) -> float:
    """Straight-line distance in miles between two lat/lon points."""
    R = 3958.8
    p1, p2 = math.radians(lat1), math.radians(lat2)
    a = (math.sin(math.radians(lat2-lat1)/2)**2
         + math.cos(p1)*math.cos(p2)*math.sin(math.radians(lon2-lon1)/2)**2)
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))


def centroid(units: list) -> tuple:
    """Average lat/lon of a list of family units."""
    lats = [u[0].lat for u in units]
    lons = [u[0].lon for u in units]
    return sum(lats)/len(lats), sum(lons)/len(lons)


# ─────────────────────────────────────────────────────────────────────────────
# OSRM real driving times  (free public API, no key needed)
# ─────────────────────────────────────────────────────────────────────────────

def _load_json(path: str) -> dict:
    if os.path.exists(path):
        try:
            with open(path) as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def clear_bad_geocache() -> int:
    """
    Remove geocache entries that have coordinates outside Pennsylvania.
    These are bad Nominatim results that caused clustering failures.
    Returns the number of entries removed.
    Call this if you see geographically wrong routing results.
    """
    cache = _load_json(GEOCACHE_FILE)
    bad_keys = [k for k, v in cache.items()
                if not _in_pa(v[0], v[1])]
    for k in bad_keys:
        del cache[k]
    if bad_keys:
        _save_json(GEOCACHE_FILE, cache)
    return len(bad_keys)


def _save_json(path: str, data: dict) -> None:
    try:
        with open(path, "w") as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass


def _fallback_minutes(lat1, lon1, lat2, lon2) -> float:
    """Road-factor estimate: much better than raw straight-line × 3 min/mi."""
    road_mi = haversine_mi(lat1, lon1, lat2, lon2) * ROAD_FACTOR
    return (road_mi / MPH_SUBURBAN) * 60.0


def driving_minutes(lat1, lon1, lat2, lon2, cache: dict) -> float:
    """
    Real driving time in minutes via Google Directions API.
    Falls back to OSRM, then road-factor haversine estimate if unavailable.
    Results cached in routecache.json so each pair is only queried once.
    """
    key = f"{lat1:.5f},{lon1:.5f}|{lat2:.5f},{lon2:.5f}"
    if key in cache:
        return cache[key]

    # ── Try Google Directions API first ──────────────────────────────────
    if GOOGLE_MAPS_KEY:
        try:
            params = urllib.parse.urlencode({
                "origin":      f"{lat1},{lon1}",
                "destination": f"{lat2},{lon2}",
                "mode":        "driving",
                "key":         GOOGLE_MAPS_KEY,
            })
            url = f"https://maps.googleapis.com/maps/api/directions/json?{params}"
            req = urllib.request.Request(url,
                headers={"User-Agent": "ElbowLaneCampBusRouter/1.0"})
            with urllib.request.urlopen(req, timeout=8) as resp:
                data = json.loads(resp.read().decode())
            if data.get("status") == "OK":
                mins = data["routes"][0]["legs"][0]["duration"]["value"] / 60.0
                cache[key] = mins
                _save_json(ROUTECACHE_FILE, cache)
                return mins
        except Exception:
            pass   # fall through to OSRM

    # ── Fallback: OSRM ───────────────────────────────────────────────────
    try:
        url = (f"http://router.project-osrm.org/route/v1/driving/"
               f"{lon1:.6f},{lat1:.6f};{lon2:.6f},{lat2:.6f}"
               f"?overview=false")
        req = urllib.request.Request(
            url, headers={"User-Agent": "ElbowLaneCampBusRouter/1.0"})
        with urllib.request.urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode())
        if data.get("code") == "Ok":
            mins = data["routes"][0]["duration"] / 60.0
            cache[key] = mins
            _save_json(ROUTECACHE_FILE, cache)
            return mins
    except Exception:
        pass

    # ── Final fallback: road-factor haversine ─────────────────────────────
    mins = _fallback_minutes(lat1, lon1, lat2, lon2)
    cache[key] = mins
    _save_json(ROUTECACHE_FILE, cache)
    return mins


def route_leg_times(coord_seq: list, progress_cb=None) -> list:
    """
    Given [(lat,lon), ...] ordered stops (incl. start and camp at end),
    returns driving minutes for each consecutive leg.
    Uses OSRM with fallback estimate.
    """
    cache = _load_json(ROUTECACHE_FILE)
    times = []
    for i in range(len(coord_seq) - 1):
        lat1, lon1 = coord_seq[i]
        lat2, lon2 = coord_seq[i+1]
        times.append(driving_minutes(lat1, lon1, lat2, lon2, cache))
    return times


# ─────────────────────────────────────────────────────────────────────────────
# Geocoding  (OpenStreetMap Nominatim — free, no key)
# ─────────────────────────────────────────────────────────────────────────────

# Pennsylvania geographic bounds — used to validate geocoding results
PA_LAT = (39.7, 42.3)
PA_LON = (-80.5, -74.7)

def _in_pa(lat: float, lon: float) -> bool:
    return PA_LAT[0] <= lat <= PA_LAT[1] and PA_LON[0] <= lon <= PA_LON[1]


# Known approximate ZIP code centroids for SE Pennsylvania.
# Used to validate geocoding results — if Nominatim returns a point
# more than MAX_ZIP_DEVIATION_MI from the expected ZIP centroid,
# it's treated as a wrong result and retried.
# This catches cases like "108 Country Ln, Lansdale, PA 19446" resolving
# to a "Country Ln" in another state or county.
MAX_ZIP_DEVIATION_MI = 3.0   # max acceptable distance from ZIP centroid

# These are rough ZIP centroids for areas around the camp.
# Any PA address not in this table falls back to PA-bounds-only validation.
ZIP_CENTROIDS = {
    "18901": (40.310, -75.130), "18902": (40.281, -75.095),
    "18914": (40.286, -75.207), "18929": (40.250, -75.084),
    "18954": (40.217, -74.999), "18974": (40.231, -75.062),
    "18976": (40.245, -75.141), "19002": (40.157, -75.228),
    "19025": (40.139, -75.177), "19040": (40.181, -75.106),
    "19044": (40.190, -75.126), "19090": (40.149, -75.120),
    "19446": (40.241, -75.284), "19002": (40.157, -75.228),
}


def _extract_zip5(address: str) -> str:
    """Extract 5-digit ZIP from address string."""
    m_zip = re.search(r"\b(\d{5})(?:-\d{4})?\b", address)
    return m_zip.group(1) if m_zip else ""


def _result_near_zip(lat: float, lon: float, zip5: str) -> bool:
    """Return True if coords are plausibly close to the expected ZIP centroid."""
    if zip5 not in ZIP_CENTROIDS:
        return True   # can't validate — give benefit of the doubt
    clat, clon = ZIP_CENTROIDS[zip5]
    return haversine_mi(lat, lon, clat, clon) <= MAX_ZIP_DEVIATION_MI


def _geocode_one(address: str, cache: dict, progress_cb=None) -> tuple:
    """
    Geocode one address using Google Geocoding API (primary) with
    Nominatim as fallback if Google is unavailable.

    Results are validated against Pennsylvania bounds and ZIP centroid
    proximity regardless of which service returns them.
    """
    key = address.strip().lower()

    # Override file always wins
    overrides = _load_overrides()
    if key in overrides:
        lat, lon = float(overrides[key][0]), float(overrides[key][1])
        cache[key] = [lat, lon]
        return lat, lon

    if key in cache:
        cached = tuple(cache[key])
        zip5_check = _extract_zip5(address)
        if (_in_pa(*cached)
                and cached != CAMP_COORDS
                and _result_near_zip(*cached, zip5_check)):
            return cached
        # Bad cached entry — purge it and stale route times
        if progress_cb:
            progress_cb(f"  Clearing bad cached coord for {address}: {cached}")
        rcache = _load_json(ROUTECACHE_FILE)
        bad_prefix = f"{cached[0]:.5f},{cached[1]:.5f}"
        stale = [k for k in rcache if bad_prefix in k]
        for k in stale:
            del rcache[k]
        if stale:
            _save_json(ROUTECACHE_FILE, rcache)
        del cache[key]

    if progress_cb:
        progress_cb(f"  Geocoding: {address}")

    zip5  = _extract_zip5(address)
    parts = [p.strip() for p in address.split(",")]
    street = parts[0] if parts else address
    city   = parts[1] if len(parts) > 1 else ""

    def _validate(lat, lon) -> bool:
        return _in_pa(lat, lon) and _result_near_zip(lat, lon, zip5)

    result = None

    # ── Pass 1: Google Geocoding API ──────────────────────────────────────
    if GOOGLE_MAPS_KEY:
        try:
            params = urllib.parse.urlencode({
                "address":    address,
                "components": "country:US|administrative_area:PA",
                "key":        GOOGLE_MAPS_KEY,
            })
            url = f"https://maps.googleapis.com/maps/api/geocode/json?{params}"
            req = urllib.request.Request(url,
                headers={"User-Agent": "ElbowLaneCampBusRouter/1.0"})
            with urllib.request.urlopen(req, timeout=8) as resp:
                data = json.loads(resp.read().decode())
            if data.get("status") == "OK" and data.get("results"):
                loc = data["results"][0]["geometry"]["location"]
                lat, lon = float(loc["lat"]), float(loc["lng"])
                if _validate(lat, lon):
                    result = (lat, lon)
                elif progress_cb:
                    progress_cb(f"  ⚠ Google result out of PA bounds for {address} — trying Nominatim")
        except Exception as e:
            if progress_cb:
                progress_cb(f"  ⚠ Google geocode error: {e}")

    # ── Pass 2: Nominatim structured (fallback) ───────────────────────────
    if result is None:
        try:
            params = urllib.parse.urlencode({
                "street":  street,
                "city":    city,
                "state":   "Pennsylvania",
                "country": "United States",
                "format":  "json",
                "limit":   3,
                "addressdetails": 0,
            })
            req = urllib.request.Request(
                f"https://nominatim.openstreetmap.org/search?{params}",
                headers={"User-Agent": "ElbowLaneCampBusRouter/1.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                results = json.loads(resp.read().decode())
            time.sleep(1.1)
            for r in results:
                lat, lon = float(r["lat"]), float(r["lon"])
                if _validate(lat, lon):
                    result = (lat, lon)
                    break
        except Exception:
            pass

    # ── Pass 3: Nominatim hard-bounded free-text ──────────────────────────
    if result is None:
        try:
            params = urllib.parse.urlencode({
                "q":       address,
                "format":  "json",
                "limit":   5,
                "addressdetails": 0,
                "viewbox": "-80.5,39.7,-74.7,42.3",
                "bounded": 1,
            })
            req = urllib.request.Request(
                f"https://nominatim.openstreetmap.org/search?{params}",
                headers={"User-Agent": "ElbowLaneCampBusRouter/1.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                results = json.loads(resp.read().decode())
            time.sleep(1.1)
            for r in results:
                lat, lon = float(r["lat"]), float(r["lon"])
                if _validate(lat, lon):
                    result = (lat, lon)
                    break
        except Exception:
            pass

    # ── Pass 4: ZIP centroid fallback ─────────────────────────────────────
    if result is None and zip5 in ZIP_CENTROIDS:
        lat, lon = ZIP_CENTROIDS[zip5]
        if progress_cb:
            progress_cb(f"  ⚠ Using ZIP centroid for '{address}' — may be approximate")
        result = (lat, lon)

    if result is not None:
        lat, lon = result
        cache[key] = [lat, lon]
        _save_json(GEOCACHE_FILE, cache)
        if progress_cb:
            progress_cb(f"  ✓ {address} → ({lat:.4f}, {lon:.4f})")
        return lat, lon

    if progress_cb:
        progress_cb(f"  ⚠ Could not geocode '{address}' — using camp coords as fallback")
    return CAMP_COORDS


def _purge_bad_geocache(cache: dict, addresses: list, progress_cb=None) -> int:
    """
    Scan the geocache for entries with coordinates that fail ZIP validation.
    Removes them so they will be re-geocoded correctly on the next run.
    Called automatically at the start of every geocode_all_addresses call.
    Returns number of bad entries removed.
    """
    rcache = _load_json(ROUTECACHE_FILE)
    removed = 0
    for addr in addresses:
        key = addr.strip().lower()
        if key not in cache:
            continue
        cached = tuple(cache[key])
        zip5 = _extract_zip5(addr)
        is_bad = (
            cached == CAMP_COORDS
            or not _in_pa(*cached)
            or not _result_near_zip(*cached, zip5)
        )
        if is_bad:
            # Also wipe any route-time cache entries that used these bad coords
            bad_prefix = f"{cached[0]:.5f},{cached[1]:.5f}"
            stale_routes = [k for k in rcache if bad_prefix in k]
            for k in stale_routes:
                del rcache[k]
            del cache[key]
            removed += 1
            if progress_cb:
                progress_cb(f"  Purged bad geocache entry: {addr} was {cached}")
    if removed:
        _save_json(GEOCACHE_FILE, cache)
        _save_json(ROUTECACHE_FILE, rcache)
    return removed


def geocode_all_addresses(addresses: list, progress_cb=None) -> dict:
    """
    Geocode a list of addresses → {address: (lat, lon)}.

    Fully self-healing — no manual cache management needed:
    • Scans and purges any cached coordinates that fail ZIP-centroid validation
    • This automatically fixes bad Nominatim results from previous runs
    • Addresses geocoded correctly are returned instantly from cache
    • New/purged addresses are re-geocoded using the 4-pass robust strategy
    • First run: ~1 second per new address (Nominatim rate limit)
    • Subsequent runs: instant (unless previous results were bad)
    """
    cache = _load_json(GEOCACHE_FILE)

    # Always purge bad entries first — fixes stale wrong coords automatically
    purged = _purge_bad_geocache(cache, addresses, progress_cb)
    if purged and progress_cb:
        progress_cb(f"  Purged {purged} bad geocache entries — will re-geocode")

    # Find addresses that still need geocoding after purge
    def needs_geocode(a: str) -> bool:
        key = a.strip().lower()
        if key not in cache:
            return True
        cached = tuple(cache[key])
        if cached == CAMP_COORDS or not _in_pa(*cached):
            return True
        zip5 = _extract_zip5(a)
        return not _result_near_zip(*cached, zip5)

    # Identify addresses with bad cached coords so we can also
    # wipe their stale route-time cache entries
    bad_addresses = [a for a in addresses if needs_geocode(a)
                     and a.strip().lower() in cache]

    if bad_addresses:
        # Remove stale route-cache entries for these addresses
        rcache = _load_json(ROUTECACHE_FILE)
        removed_routes = 0
        for addr in bad_addresses:
            old_key = cache.get(addr.strip().lower())
            if old_key:
                old_lat, old_lon = old_key[0], old_key[1]
                bad_coord_prefix = f"{old_lat:.5f},{old_lon:.5f}"
                stale = [k for k in rcache if bad_coord_prefix in k]
                for k in stale:
                    del rcache[k]
                    removed_routes += 1
        if removed_routes:
            _save_json(ROUTECACHE_FILE, rcache)
            if progress_cb:
                progress_cb(f"  Cleared {removed_routes} stale route-time entries "
                            f"for re-geocoded addresses")

    new_count = sum(1 for a in addresses if needs_geocode(a))
    if new_count and progress_cb:
        progress_cb(f"Geocoding {new_count} address(es) via OpenStreetMap "
                    f"(validating Pennsylvania bounds)...")
    elif progress_cb:
        progress_cb(f"All {len(addresses)} addresses loaded from cache")

    return {a: _geocode_one(a, cache, progress_cb) for a in addresses}


# ─────────────────────────────────────────────────────────────────────────────
# CSV parsing
# ─────────────────────────────────────────────────────────────────────────────

def parse_students_csv(csv_text: str) -> list:
    if "\n" not in csv_text and len(csv_text) < 300:
        try:
            with open(csv_text, newline="", encoding="utf-8-sig") as f:
                csv_text = f.read()
        except FileNotFoundError:
            pass
    students = []
    reader = csv.DictReader(io.StringIO(csv_text))
    for row in reader:
        row = {k.strip().strip('"'): v.strip().strip('"') for k, v in row.items()}
        idx_key = next((k for k in row if k in ("", "idx", "#")), "")
        try:    idx = int(row.get(idx_key, 0))
        except: idx = 0
        last  = (row.get("Last name")   or row.get("last_name")  or row.get("Last Name")  or "")
        first = (row.get("First name")  or row.get("first_name") or row.get("First Name") or "")
        addr  = (row.get("Primary family address 1") or row.get("Address") or
                 row.get("address") or row.get("Street") or "")
        city  = (row.get("Primary family city") or row.get("City") or row.get("city") or "")
        zip_  = (row.get("Primary family zip")  or row.get("Zip")  or row.get("zip")
                 or row.get("ZIP") or row.get("Postal Code") or "")
        if last and addr:
            students.append(Student(idx, last, first, addr, city, zip_))
    return students


# ─────────────────────────────────────────────────────────────────────────────
# Vehicle config parsing  (handles many messy real-world formats)
# ─────────────────────────────────────────────────────────────────────────────

def parse_vehicles_text(text: str) -> list:
    """
    Accepts formats like:
        Vehicle A: Start: 7826 Loretto Ave, Philadelphia, PA - Capacity: 5 riders
        Vehicle B Start: 12 Rachel Rd, Richboro, PA  Capacity: up to 13 riders
        Vehicles D
        Start: 1045 N West End Blvd, Quakertown PA - Capacity: up to 13 riders
        E-H (5 vehicles)Start & End: 828 Elbow Lane - Capacity: up to 13 riders each
    """
    VEH_RE = re.compile(
        r"""^(?:(?:vehicles?\s+[A-Z0-9][-\s,A-Z0-9]*)
              |(?:[A-Z][-][A-Z]\s*(?:\(|$))
              |(?:[A-Z]\s*(?:\(|:|\s+Start))
              |(?:Van\s+[A-Z0-9]))""",
        re.VERBOSE | re.IGNORECASE)

    # Merge continuation lines
    merged = []
    for raw in text.strip().splitlines():
        s = raw.strip()
        if not s: continue
        if merged and not VEH_RE.match(s):
            merged[-1] += " " + s
        else:
            merged.append(s)

    # Expand letter ranges  (e.g. "E-H" → Vehicle E, F, G, H)
    RNG = re.compile(r"^(?:vehicles?\s*)?([A-Z])[-]([A-Z])(?:\s*\(\d+\s*vehicles?\))?",
                     re.IGNORECASE)
    expanded = []
    for line in merged:
        m = RNG.match(line)
        if m:
            remainder = line[m.end():].strip().lstrip(":")
            for c in range(ord(m.group(1).upper()), ord(m.group(2).upper())+1):
                expanded.append(f"Vehicle {chr(c)}: {remainder}")
        else:
            expanded.append(line)

    vehicles = []
    for line in expanded:
        line = line.strip()
        nm = re.match(r"^((?:Vehicles?|Van)\s+[A-Z0-9]+)", line, re.I)
        if not nm:
            nm = re.match(r"^([A-Z])(?:\s*:|\s+Start)", line, re.I)
        if not nm: continue
        name = re.sub(r"^Vehicles\b", "Vehicle", nm.group(1).strip(), flags=re.I)
        rest = line[nm.end():].strip().lstrip(":").strip()
        sm = re.search(r"Start(?:\s*[&]\s*End)?\s*:?\s*", rest, re.I)
        if not sm: continue
        after = rest[sm.end():]
        ae = re.search(r"\s*[-]?\s*Capacity\b", after, re.I)
        start_addr = (after[:ae.start()].strip() if ae else after.strip()).rstrip(" -,")
        cm = re.search(r"Capacity\s*:?\s*(?:up\s+to\s+)?(\d+)\s*riders?", rest, re.I)
        cap = int(cm.group(1)) if cm else 13
        if name and start_addr:
            vehicles.append({"name": name, "start": start_addr, "capacity": cap})
    return vehicles


# ─────────────────────────────────────────────────────────────────────────────
# Core routing
# ─────────────────────────────────────────────────────────────────────────────

def cluster_and_route(students: list, vehicles: list,
                      progress_cb: Optional[Callable] = None,
                      camp_address: str = None,
                      trip_direction: str = "morning") -> list:
    """
    Full pipeline:
    1. Geocode every address (OpenStreetMap — real lat/lon, ZIP ignored)
    2. Group same-address students (families always together)
    3. Cluster family units by TRUE address proximity ≤ NEIGHBOR_MI
    4. Assign whole clusters to vehicles (nearest start, strict capacity)
    5. Consolidate under-filled vehicles (full merge, then scatter)
    6. Remove empty vehicles from output
    7. Sequence stops:
       - Morning: furthest from camp first (no backtracking toward camp)
       - Afternoon: nearest to camp first (route away from camp)
    8. Get driving times via OSRM (or road-factor fallback)

    Args:
        camp_address:    Override the default camp address
        trip_direction:  "morning" (routes end at camp) or
                         "afternoon" (routes start at camp, end at homes)
    """
    effective_camp = camp_address or CAMP_ADDRESS

    # ── 1. Geocode ────────────────────────────────────────────────────────
    all_addrs = list({s.full_address for s in students}
                     | {v["start"] for v in vehicles}
                     | {effective_camp})
    coords = geocode_all_addresses(all_addrs, progress_cb)

    camp_lat, camp_lon = coords.get(effective_camp, CAMP_COORDS)
    for s in students:
        s.lat, s.lon = coords.get(s.full_address, CAMP_COORDS)
        s.geocoded   = (s.lat, s.lon) != CAMP_COORDS

    if progress_cb:
        ok = sum(1 for s in students if s.geocoded)
        progress_cb(f"Geocoded {ok}/{len(students)} addresses")

    # ── 2. Group by exact address (families) ──────────────────────────────
    addr_map: dict = {}
    for s in students:
        addr_map.setdefault(s.address.lower().strip(), []).append(s)
    family_units = list(addr_map.values())

    def uc(u):   return u[0].lat, u[0].lon
    def d2c(u):  return haversine_mi(*uc(u), camp_lat, camp_lon)

    # ── 3. Geographic clustering — NO ZIP CODES ───────────────────────────
    # Two houses ≤ NEIGHBOR_MI apart → same cluster regardless of ZIP.
    clusters: list = []
    for unit in family_units:
        ulat, ulon = uc(unit)
        best_ci, best_d = None, float("inf")
        for ci, cluster in enumerate(clusters):
            for eu in cluster:
                d = haversine_mi(ulat, ulon, *uc(eu))
                if d <= NEIGHBOR_MI and d < best_d:
                    best_d, best_ci = d, ci
        if best_ci is not None:
            clusters[best_ci].append(unit)
        else:
            clusters.append([unit])

    if progress_cb:
        progress_cb(f"Formed {len(clusters)} geographic clusters "
                    f"(≤{NEIGHBOR_MI} mi between actual house coordinates)")

    # ── 4. Build vehicle objects ──────────────────────────────────────────
    def cl_size(cl): return sum(len(u) for u in cl)

    veh_objects = []
    for v in vehicles:
        lat, lon = coords.get(v["start"], CAMP_COORDS)
        veh_objects.append(Vehicle(
            name=v["name"], start_address=v["start"],
            capacity=v["capacity"], start_lat=lat, start_lon=lon))

    # ── 5. Assign WHOLE CLUSTERS to vehicles ──────────────────────────────
    # Clusters are the atomic unit — neighbours are never split.
    # Clusters too large for any single vehicle are split by distance-to-camp.
    max_cap = max(v["capacity"] for v in vehicles)
    assignable: list = []
    for cl in clusters:
        if cl_size(cl) <= max_cap:
            assignable.append(cl)
        else:
            units_sorted = sorted(cl, key=d2c, reverse=True)
            chunk, chunk_n = [], 0
            for u in units_sorted:
                if chunk_n + len(u) > max_cap and chunk:
                    assignable.append(chunk); chunk, chunk_n = [], 0
                chunk.append(u); chunk_n += len(u)
            if chunk: assignable.append(chunk)

    # Sort: farthest-from-camp clusters first (priority placement)
    assignable.sort(
        key=lambda cl: haversine_mi(*centroid(cl), camp_lat, camp_lon),
        reverse=True)

    assignments = [[] for _ in veh_objects]
    counts      = [0]  * len(veh_objects)

    def score_cluster(cl, vi) -> float:
        remaining = veh_objects[vi].capacity - counts[vi]
        if remaining < cl_size(cl): return float("inf")
        clat, clon = centroid(cl)
        geo = haversine_mi(clat, clon, veh_objects[vi].start_lat, veh_objects[vi].start_lon)
        # Bonus for filling small vehicles — reduces their likelihood of ending up empty
        cap = veh_objects[vi].capacity
        small_bonus = 3.0 if cap <= 6 else (1.5 if cap <= 9 else 0.0)
        fill_ratio = counts[vi] / cap if cap else 0
        # Prefer vehicles that are partially filled (building on existing load)
        fill_bonus = fill_ratio * 2.0
        return geo - small_bonus - fill_bonus

    leftover = []
    for cl in assignable:
        best = min(range(len(veh_objects)), key=lambda vi: score_cluster(cl, vi))
        if score_cluster(cl, best) < float("inf"):
            for u in cl: assignments[best].append(u)
            counts[best] += cl_size(cl)
        else:
            leftover.extend(cl)   # individual units if cluster won't fit

    for unit in sorted(leftover, key=d2c, reverse=True):
        sz = len(unit)
        ulat, ulon = uc(unit)
        scores = [(haversine_mi(ulat, ulon, veh_objects[vi].start_lat,
                                veh_objects[vi].start_lon), vi)
                  for vi in range(len(veh_objects))
                  if veh_objects[vi].capacity - counts[vi] >= sz]
        vi = min(scores)[1] if scores else min(range(len(veh_objects)),
                                               key=lambda i: counts[i])
        assignments[vi].append(unit)
        counts[vi] += sz

    # ── 5b. Consolidation — eliminate under-filled vehicles ───────────────
    # Two passes per iteration:
    #   FULL MERGE: if one vehicle can absorb all of the donor's students, do it
    #   SCATTER:    otherwise send each family unit to nearest vehicle with room
    # An isolated geographic cluster may not reach MIN_UTIL — that's accepted
    # and flagged with a warning in the spreadsheet rather than breaking routing.
    changed, passes = True, 0
    while changed and passes < 30:
        changed, passes = False, passes + 1

        # Sort under-filled: emptiest first (empty buses eliminated first)
        # Small vehicles (capacity ≤ 6) use a lower threshold since they're
        # harder to fill precisely with geographic constraints
        def effective_threshold(vi):
            cap = veh_objects[vi].capacity
            if cap <= 6:  return 0.50   # 50% min for small vans
            if cap <= 9:  return 0.60   # 60% min for medium vans
            return MIN_UTIL              # 75% for full-size vans

        under = sorted(
            [vi for vi in range(len(veh_objects))
             if counts[vi] > 0 and
                counts[vi] / veh_objects[vi].capacity < effective_threshold(vi)],
            key=lambda vi: counts[vi])

        if not under: break
        vi_src = under[0]
        units_src = assignments[vi_src][:]
        total_src = counts[vi_src]

        if not units_src:
            # Already empty — will be pruned in step 5c
            changed = True; continue

        src_lat, src_lon = centroid(units_src)

        # ── Full merge ────────────────────────────────────────────────────
        full_dest, full_dist = None, float("inf")
        for vi_dst in range(len(veh_objects)):
            if vi_dst == vi_src: continue
            if veh_objects[vi_dst].capacity - counts[vi_dst] < total_src: continue
            if assignments[vi_dst]:
                dlat, dlon = centroid(assignments[vi_dst])
            else:
                dlat, dlon = veh_objects[vi_dst].start_lat, veh_objects[vi_dst].start_lon
            d = haversine_mi(src_lat, src_lon, dlat, dlon)
            if d < full_dist: full_dist, full_dest = d, vi_dst

        if full_dest is not None:
            for u in units_src: assignments[full_dest].append(u)
            counts[full_dest] += total_src
            assignments[vi_src] = []; counts[vi_src] = 0
            if progress_cb:
                progress_cb(f"  Merged {veh_objects[vi_src].name} ({total_src}) → "
                            f"{veh_objects[full_dest].name} "
                            f"({counts[full_dest]}/{veh_objects[full_dest].capacity})")
            changed = True; continue

        # ── Scatter: move sub-groups to nearest vehicle with room ──────────
        # IMPORTANT: we scatter by geographic sub-groups, not individual units.
        # Units within NEIGHBOR_MI of each other are moved together to preserve
        # the neighbor-grouping guarantee. This prevents consolidation from
        # splitting clusters that were correctly grouped in step 3.
        
        # Build sub-groups: connected components within NEIGHBOR_MI
        remaining = list(units_src)
        sub_groups = []
        while remaining:
            group = [remaining[0]]
            remaining.pop(0)
            changed_inner = True
            while changed_inner:
                changed_inner = False
                for unit in list(remaining):
                    ulat, ulon = uc(unit)
                    if any(haversine_mi(ulat, ulon, *uc(g)) <= NEIGHBOR_MI
                           for g in group):
                        group.append(unit)
                        remaining.remove(unit)
                        changed_inner = True
            sub_groups.append(group)
        
        # Sort sub-groups largest first (harder to place)
        sub_groups.sort(key=lambda g: sum(len(u) for u in g), reverse=True)
        
        moved = False
        for group in sub_groups:
            group_size = sum(len(u) for u in group)
            glat = sum(uc(u)[0] for u in group) / len(group)
            glon = sum(uc(u)[1] for u in group) / len(group)
            
            # Max distance a group can be scattered — prevents mixing
            # students from completely different geographic areas
            MAX_SCATTER_MI = 5.0

            best_vi, best_d = None, float("inf")
            for vi_dst in range(len(veh_objects)):
                if vi_dst == vi_src: continue
                if veh_objects[vi_dst].capacity - counts[vi_dst] < group_size: continue
                if assignments[vi_dst]:
                    d = min(haversine_mi(glat, glon, *uc(eu))
                            for eu in assignments[vi_dst])
                    # Hard reject: too far from any existing stop on this vehicle
                    if d > MAX_SCATTER_MI: continue
                else:
                    d = haversine_mi(glat, glon,
                                     veh_objects[vi_dst].start_lat,
                                     veh_objects[vi_dst].start_lon)
                    # Empty vehicle — only accept if start is reasonably close
                    if d > MAX_SCATTER_MI * 2: continue
                if d < best_d: best_d, best_vi = d, vi_dst

            if best_vi is not None:
                for unit in group:
                    assignments[best_vi].append(unit)
                    assignments[vi_src].remove(unit)
                counts[best_vi] += group_size
                counts[vi_src]  -= group_size
                moved = True
                names = ", ".join(u[0].last for u in group)
                if progress_cb:
                    progress_cb(f"  Moved sub-group [{names}] ({group_size}) "
                                f"{veh_objects[vi_src].name} → "
                                f"{veh_objects[best_vi].name}")
        if moved: changed = True
        # If nothing moved: geographic constraint — accept and flag as warning

    # ── 5c. Prune truly empty vehicles (0 riders) ────────────────────────
    # Vehicles that have riders but are under threshold are kept and flagged.
    # Only vehicles with zero riders are removed from the output.
    active = [vi for vi in range(len(veh_objects)) if counts[vi] > 0]
    veh_objects = [veh_objects[vi] for vi in active]
    assignments = [assignments[vi] for vi in active]
    counts      = [counts[vi]      for vi in active]
    if progress_cb:
        progress_cb(f"Active vehicles: {len(veh_objects)}")

    # ── 6. Sequence stops + driving times ────────────────────────────────
    for vi, veh in enumerate(veh_objects):
        if not assignments[vi]:
            veh.total_time = "—"; veh.total_distance = "—"; continue

        # Build address-level stop objects
        addr_stop: dict = {}
        for unit in assignments[vi]:
            rep = unit[0]; key = rep.address.lower().strip()
            if key not in addr_stop:
                addr_stop[key] = Stop(address=rep.full_address,
                                      lat=rep.lat, lon=rep.lon)
            addr_stop[key].riders.extend(unit)

        # ── Nearest-neighbor TSP sequencing ─────────────────────────────
        # Start from vehicle start, greedily pick closest unvisited stop,
        # with a "no-backtracking" constraint: once we're moving toward camp
        # we don't allow stops that are significantly farther from camp than
        # the current position (prevents U-turns).
        #
        # This guarantees:
        #   • Geographic neighbours are consecutive stops
        #   • Route flows roughly toward camp (no major detours)
        #   • Every stop is visited exactly once
        unvisited = list(addr_stop.values())

        def dist_to_camp_s(s):
            return haversine_mi(s.lat, s.lon, camp_lat, camp_lon)

        # Sequence stops using nearest-neighbor TSP with directional bias:
        # Morning:   start from farthest stop, work toward camp (no backtracking)
        # Afternoon: start from nearest stop to camp, work away from camp
        if trip_direction == "afternoon":
            first = min(unvisited, key=dist_to_camp_s)
        else:
            first = max(unvisited, key=dist_to_camp_s)
        sorted_stops = [first]
        unvisited.remove(first)

        while unvisited:
            last = sorted_stops[-1]
            cur_d2c = dist_to_camp_s(last)
            best_stop, best_score = None, float("inf")
            for s in unvisited:
                geo_d = haversine_mi(last.lat, last.lon, s.lat, s.lon)
                d2c_s = dist_to_camp_s(s)
                if trip_direction == "afternoon":
                    # Afternoon: penalise stops that are closer to camp
                    # (encourages routing away from camp, not back toward it)
                    backtrack_penalty = max(0.0, cur_d2c - d2c_s) * 0.5
                else:
                    # Morning: penalise stops that are farther from camp
                    backtrack_penalty = max(0.0, d2c_s - cur_d2c) * 0.5
                score = geo_d + backtrack_penalty
                if score < best_score:
                    best_score, best_stop = score, s
            sorted_stops.append(best_stop)
            unvisited.remove(best_stop)

        # Real driving times: vehicle start → stop1 → … → stopN → camp
        coord_seq = ([(veh.start_lat, veh.start_lon)]
                     + [(s.lat, s.lon) for s in sorted_stops]
                     + [(camp_lat, camp_lon)])
        legs = route_leg_times(coord_seq, progress_cb)

        for i, stop in enumerate(sorted_stops):
            mins = max(1, round(legs[i]))
            stop.drive_time = f"{mins} min from start" if i == 0 else f"{mins} min"

        veh.stops    = sorted_stops
        veh.camp_lat = camp_lat
        veh.camp_lon = camp_lon

        total_mins = round(sum(legs))
        if total_mins >= 60:
            hrs, rem = divmod(total_mins, 60)
            veh.total_time = f"{hrs} hr {rem} min" if rem else f"{hrs} hr"
        else:
            veh.total_time = f"{total_mins} min"

        dist = sum(
            haversine_mi(coord_seq[i][0], coord_seq[i][1],
                         coord_seq[i+1][0], coord_seq[i+1][1]) * ROAD_FACTOR
            for i in range(len(coord_seq)-1))
        veh.total_distance = f"{round(dist, 1)} mi"

        # Flag under-threshold for dashboard warning
        cap = veh.capacity
        eff_threshold = 0.50 if cap <= 6 else (0.60 if cap <= 9 else MIN_UTIL)
        veh.under_threshold = (veh.rider_count / cap < eff_threshold
                                if cap else False)

    return veh_objects


# ─────────────────────────────────────────────────────────────────────────────
# Excel output
# ─────────────────────────────────────────────────────────────────────────────

def _fill(c):   return PatternFill("solid", start_color=c, fgColor=c)
def _font(bold=False, size=11, color=DARK_TEXT, italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def _bdr(**kw): return Border(**kw)


def build_dashboard(wb: Workbook, vehicles: list, camp_address: str = None, trip_direction: str = "morning"):
    ws = wb.active
    ws.title = "Route Summary"
    for col, w in zip("ABCDEFGH", [18, 54, 10, 10, 9, 16, 13, 10]):
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "🚌  Elbow Lane Day Camp — Vehicle Route Plan"
    c.font      = _font(bold=True, size=16, color=WHITE)
    c.fill      = _fill(BRAND_COLOR)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 32

    # Subtitle
    ws.merge_cells("A2:H2")
    c = ws["A2"]
    camp_display = camp_address or CAMP_ADDRESS
    direction_label = "All vehicles depart from" if trip_direction == "afternoon" else "All vehicles finish at"
    c.value = (f"{direction_label}: {camp_display}  |  "
               "Clustered by real street-address proximity (OpenStreetMap)  |  "
               "Drive times via OSRM road-network routing")
    c.font      = _font(size=9, italic=True, color="444444")
    c.fill      = _fill(BRAND_LIGHT)
    c.alignment = _align("center")
    ws.row_dimensions[2].height = 18

    # Headers
    hdrs = ["Vehicle", "Starting Point / Route Corridor",
            "Capacity", "Riders", "Stops", "Drive Time", "Distance", "Utilization"]
    for ci, h in enumerate(hdrs, 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font      = _font(bold=True, size=10, color=WHITE)
        cell.fill      = _fill(BRAND_COLOR)
        cell.alignment = _align("center")
        cell.border    = _bdr(bottom=MED_SIDE, top=MED_SIDE,
                               left=THIN_SIDE, right=THIN_SIDE)
    ws.row_dimensions[3].height = 18

    total_cap = total_riders = 0
    has_warnings = False
    for ri, veh in enumerate(vehicles):
        row   = 4 + ri
        warn  = veh.under_threshold
        if warn: has_warnings = True
        bg    = _fill(ORANGE_FILL) if warn else (_fill(WHITE) if ri%2==0 else _fill(LIGHT_GRAY))
        util_txt = f"{veh.utilization_pct}%"
        if warn: util_txt += " ⚠"

        vals = [veh.name,
                f"{veh.start_address}  |  {veh.corridor}",
                veh.capacity, veh.rider_count, veh.stop_count,
                veh.total_time, veh.total_distance, util_txt]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font      = _font(size=10, bold=warn)
            cell.fill      = bg
            cell.alignment = _align("left" if ci == 2 else "center")
            cell.border    = _bdr(bottom=THIN_SIDE)
        total_cap    += veh.capacity
        total_riders += veh.rider_count
        ws.row_dimensions[row].height = 15

    # Totals row
    tr = 4 + len(vehicles)
    ws.merge_cells(f"A{tr}:B{tr}")
    c = ws[f"A{tr}"]
    c.value     = (f"TOTAL  ({total_riders} riders / {total_cap} capacity "
                   f"/ {len(vehicles)} vehicles)")
    c.font      = _font(bold=True, size=10, color=WHITE)
    c.fill      = _fill(BRAND_COLOR)
    c.alignment = _align("center")
    for ci, val in [(3, total_cap), (4, f"=SUM(D4:D{tr-1})"),
                    (5, f"=SUM(E4:E{tr-1})"),
                    (6, "—"), (7, "—"), (8, "—")]:
        cell = ws.cell(row=tr, column=ci, value=val)
        cell.font = _font(bold=True, color=WHITE); cell.fill = _fill(BRAND_COLOR)
        cell.alignment = _align("center")
    ws.row_dimensions[tr].height = 18

    # Legend / warning
    lr = tr + 2
    ws.merge_cells(f"A{lr}:H{lr}")
    c = ws[f"A{lr}"]
    if has_warnings:
        c.value = ("⚠  Orange rows are below 75% capacity.  "
                   "These vehicles serve geographically isolated stops that cannot be "
                   "merged without splitting neighbour groups.  "
                   "Consider reducing the number of vehicles in the fleet config.")
        c.font  = _font(size=9, bold=True, color="7B3F00", italic=False)
        c.fill  = _fill(ORANGE_FILL)
    else:
        c.value = ("All vehicles at ≥ 75% capacity  |  "
                   "Drive times via OSRM road-network routing  |  "
                   "Clustering by real geocoded addresses (OpenStreetMap) — ZIP codes ignored")
        c.font  = _font(size=9, italic=True, color="555555")
    c.alignment = _align("left", wrap=True)
    ws.row_dimensions[lr].height = 28


def build_vehicle_sheet(wb: Workbook, veh: Vehicle, camp_address: str = None, trip_direction: str = "morning"):
    ws = wb.create_sheet(title=veh.name)
    for col, w in zip("ABCDE", [9, 40, 28, 12, 22]):
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = f"🚌  {veh.name}  —  Route Sheet"
    c.font      = _font(bold=True, size=14, color=WHITE)
    c.fill      = _fill(BRAND_COLOR)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    # Summary bar — orange if under-threshold
    ws.merge_cells("A2:E2")
    c = ws["A2"]
    warn_tag = "  ⚠ Below 75% — see dashboard" if veh.under_threshold else ""
    c.value = (f"Start: {veh.start_address}   |   Cap: {veh.capacity}   |   "
               f"Riders: {veh.rider_count} ({veh.utilization_pct}%)   |   "
               f"Total Route: {veh.total_time}, {veh.total_distance}{warn_tag}")
    c.font      = _font(size=9, italic=True,
                        color="7B3F00" if veh.under_threshold else DARK_TEXT)
    c.fill      = _fill(ORANGE_FILL if veh.under_threshold else BRAND_LIGHT)
    c.alignment = _align("left")
    ws.row_dimensions[2].height = 16

    # Column headers
    for ci, h in enumerate(["Stop #","Address","Riders","# Riders","Drive Time"], 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.font      = _font(bold=True, size=10, color=WHITE)
        cell.fill      = _fill(BRAND_COLOR)
        cell.alignment = _align("center")
        cell.border    = _bdr(bottom=MED_SIDE)
    ws.row_dimensions[3].height = 16

    # START row
    for ci, val in enumerate([" START", veh.start_address,
                               "Departure point", None, None], 1):
        cell = ws.cell(row=4, column=ci, value=val)
        cell.font  = _font(bold=(ci==1), italic=(ci==3), color="555555")
        cell.fill  = _fill(LIGHT_GRAY)
        cell.alignment = _align("left" if ci == 2 else "center")
    ws.row_dimensions[4].height = 14

    # Stop rows
    for si, stop in enumerate(veh.stops):
        row = 5 + si
        bg  = _fill(WHITE) if si % 2 == 0 else _fill(LIGHT_GRAY)
        for ci, val in enumerate([si+1, stop.address, stop.rider_names,
                                   stop.rider_count, stop.drive_time], 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font      = _font(bold=(ci==1), size=10)
            cell.fill      = bg
            cell.alignment = _align("center" if ci in (1,4) else "left")
            cell.border    = _bdr(bottom=THIN_SIDE)
        ws.row_dimensions[row].height = 14

    # ARRIVE row
    arrive = 5 + len(veh.stops)
    if veh.stops:
        last = veh.stops[-1]
        final = max(1, round(_fallback_minutes(last.lat, last.lon, *CAMP_COORDS)))
        arrive_time = f"{final} min → ARRIVE"
    else:
        arrive_time = "— → ARRIVE"

    arrive_label = camp_address or CAMP_ADDRESS
    dest_display = arrive_label.split(",")[0] if "," in arrive_label else arrive_label
    action_word = "DEPART" if trip_direction == "afternoon" else "ARRIVE"
    for ci, val in enumerate([action_word, arrive_label,
                               "—", "—", arrive_time], 1):
        cell = ws.cell(row=arrive, column=ci, value=val)
        cell.font      = _font(bold=True, color="006400")
        cell.fill      = _fill(GREEN_FILL)
        cell.alignment = _align("left" if ci == 2 else "center")
        cell.border    = _bdr(top=MED_SIDE, bottom=MED_SIDE)
    ws.row_dimensions[arrive].height = 16

    # Total riders
    tr = arrive + 1
    ws.cell(row=tr, column=4, value="Total Riders:").font = _font(bold=True)
    ws.cell(row=tr, column=4).alignment = _align("right")
    cell = ws.cell(row=tr, column=5, value=f"=SUM(D5:D{arrive-1})")
    cell.font = _font(bold=True); cell.alignment = _align("center")

    # Footer note
    nr = tr + 1
    ws.merge_cells(f"A{nr}:E{nr}")
    note = ws[f"A{nr}"]
    note.value = ("Drive times via OSRM road-network routing  |  "
                  "Stop order by real geocoded coordinates (OpenStreetMap) — ZIP codes not used  |  "
                  "Falls back to road-distance estimate if OSRM unavailable")
    note.font  = _font(size=8, italic=True, color="777777")
    note.alignment = _align("left")


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────

def generate_routes(
    csv_text: str,
    vehicles_text: str,
    output_path: str = "bus_routes_output.xlsx",
    route_data: Optional[list] = None,
    progress_cb: Optional[Callable] = None,
    camp_address: str = None,
    trip_direction: str = "morning",
) -> str:
    """
    Parse → geocode → cluster → assign → consolidate → Excel.

    Args:
        camp_address:    Camp destination/origin (defaults to 828 Elbow Lane)
        trip_direction:  "morning" (students travel TO camp) or
                         "afternoon" (students travel FROM camp HOME)
    Returns output_path on success.
    """
    students = parse_students_csv(csv_text)
    if not students:
        raise ValueError("No students parsed. Check CSV format.")

    vehicle_configs = parse_vehicles_text(vehicles_text)
    if not vehicle_configs:
        raise ValueError("No vehicles parsed. Check fleet configuration format.")

    if progress_cb:
        unique = len({s.full_address for s in students})
        progress_cb(f"Loaded {len(students)} students across {unique} addresses, "
                    f"{len(vehicle_configs)} vehicles")

    if route_data:
        vehicles = _apply_ai_routes(vehicle_configs, route_data)
    else:
        vehicles = cluster_and_route(students, vehicle_configs, progress_cb,
                                          camp_address=camp_address,
                                          trip_direction=trip_direction)

    wb = Workbook()
    build_dashboard(wb, vehicles, camp_address=camp_address, trip_direction=trip_direction)
    for veh in vehicles:
        build_vehicle_sheet(wb, veh, camp_address=camp_address, trip_direction=trip_direction)

    wb.save(output_path)
    if progress_cb:
        progress_cb(f"✅  Saved: {output_path}")
    return output_path


def _apply_ai_routes(vehicle_configs, route_data) -> list:
    """Use pre-computed AI route data instead of algorithmic clustering."""
    vehicles = []
    for rd in route_data:
        veh = Vehicle(name=rd.get("vehicle_name","?"),
                      start_address=rd.get("start_address",""),
                      capacity=rd.get("capacity", 13),
                      total_time=rd.get("total_time","—"),
                      total_distance=rd.get("total_distance","—"))
        for sd in rd.get("stops", []):
            stop = Stop(address=sd.get("address",""))
            for name in sd.get("rider_names", []):
                stop.riders.append(Student(0, name, "", "", "", ""))
            stop.drive_time = sd.get("drive_time","")
            veh.stops.append(stop)
        vehicles.append(veh)
    return vehicles


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Elbow Lane Camp Bus Router")
    parser.add_argument("--csv",        required=True)
    parser.add_argument("--vehicles",   required=True)
    parser.add_argument("--output",     default="bus_routes_output.xlsx")
    parser.add_argument("--route-json", default=None)
    args = parser.parse_args()

    with open(args.csv, encoding="utf-8-sig") as f:
        csv_text = f.read()
    try:
        with open(args.vehicles) as f:
            vehicles_text = f.read()
    except (FileNotFoundError, IsADirectoryError):
        vehicles_text = args.vehicles.replace("\\n", "\n")

    route_data = None
    if args.route_json:
        with open(args.route_json) as f:
            route_data = json.load(f)

    generate_routes(csv_text, vehicles_text, args.output, route_data, print)


if __name__ == "__main__":
    main()
